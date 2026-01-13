import io
import os
import logging
import re
from datetime import datetime
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import Response, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl

app = FastAPI(title="Excel Generator Service")

# Configurar CORS para permitir requests desde web y móvil
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En producción, especifica los orígenes permitidos
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ROOT = os.path.dirname(__file__)
# Buscar plantillas en la carpeta del servicio primero, luego en la raíz del proyecto
TEMPLATES_DIR = os.path.join(ROOT, "assets", "templates")
PROJECT_ROOT = os.path.dirname(os.path.dirname(ROOT))  # Subir dos niveles desde excel_generator_service
PROJECT_ASSETS_DIR = os.path.join(PROJECT_ROOT, "assets")

# Jumpers: buscar primero en templates del servicio, luego en assets del proyecto
TEMPLATE_PATH_JUMPERS = (
    os.path.join(TEMPLATES_DIR, "plantilla_jumpers.xlsx")
    if os.path.exists(os.path.join(TEMPLATES_DIR, "plantilla_jumpers.xlsx"))
    else (
        os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_jumpers.xlsx")
        if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_jumpers.xlsx"))
        else os.path.join(PROJECT_ASSETS_DIR, "plantilla_jumpers.xlsx")
    )
)
# Computo: buscar en assets/templates primero, luego en assets directamente
TEMPLATE_PATH_COMPUTO = (
    os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_inventario_computo.xlsx")
    if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_inventario_computo.xlsx"))
    else (
        os.path.join(PROJECT_ASSETS_DIR, "plantilla_inventario_computo.xlsx")
        if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "plantilla_inventario_computo.xlsx"))
        else os.path.join(TEMPLATES_DIR, "plantilla_inventario_computo.xlsx")
    )
)
# SDR: buscar primero en templates del servicio, luego en assets del proyecto
TEMPLATE_PATH_SDR = (
    os.path.join(TEMPLATES_DIR, "plantilla_sdr.xlsx")
    if os.path.exists(os.path.join(TEMPLATES_DIR, "plantilla_sdr.xlsx"))
    else os.path.join(PROJECT_ASSETS_DIR, "plantilla_SDR.xlsx")
)
# SICOR: buscar primero en templates del servicio, luego en assets del proyecto
TEMPLATE_PATH_SICOR = (
    os.path.join(TEMPLATES_DIR, "plantilla_sicor.xlsx")
    if os.path.exists(os.path.join(TEMPLATES_DIR, "plantilla_sicor.xlsx"))
    else (
        os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_sicor.xlsx")
        if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_sicor.xlsx"))
        else os.path.join(PROJECT_ASSETS_DIR, "plantilla_sicor.xlsx")
    )
)
# Bitácora: buscar primero en templates del servicio (nota: el archivo se llama platilla_bitacora.xlsx)
TEMPLATE_PATH_BITACORA = (
    os.path.join(TEMPLATES_DIR, "platilla_bitacora.xlsx")
    if os.path.exists(os.path.join(TEMPLATES_DIR, "platilla_bitacora.xlsx"))
    else (
        os.path.join(TEMPLATES_DIR, "plantilla_bitacora.xlsx")
        if os.path.exists(os.path.join(TEMPLATES_DIR, "plantilla_bitacora.xlsx"))
        else (
            os.path.join(PROJECT_ASSETS_DIR, "templates", "platilla_bitacora.xlsx")
            if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "templates", "platilla_bitacora.xlsx"))
            else os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_bitacora.xlsx")
            if os.path.exists(os.path.join(PROJECT_ASSETS_DIR, "templates", "plantilla_bitacora.xlsx"))
            else os.path.join(PROJECT_ASSETS_DIR, "platilla_bitacora.xlsx")
        )
    )
)

LAST_GENERATED_FILE_CONTENT: bytes | None = None
LAST_GENERATED_FILENAME: str | None = None

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _ensure_template(path: str):
    """Verifica que la plantilla exista, si no, crea una estructura básica"""
    if not os.path.exists(path):
        logger.warning(f"Template not found: {path}, creating basic structure")
        # No lanzamos error, crearemos la estructura básica
        return False
    return True


def _safe_set_cell_value(ws, row: int, col: int, value: Any):
    """Escribe un valor en una celda de forma segura, evitando celdas combinadas"""
    try:
        cell = ws.cell(row=row, column=col)
        
        # Verificar si la celda está en un rango combinado
        # Si la celda es parte de un merge, solo escribir en la celda principal (top-left)
        import re
        for merged_range in list(ws.merged_cells.ranges):
            range_str = str(merged_range)
            # El rango tiene formato como "A1:B2"
            if ':' in range_str:
                range_parts = range_str.split(':')
                if len(range_parts) >= 2:
                    top_left = range_parts[0]  # Ej: "A1"
                    bottom_right = range_parts[1]  # Ej: "B2"
                    
                    # Extraer fila y columna de la celda principal (top-left)
                    match_top = re.match(r'([A-Z]+)(\d+)', top_left)
                    match_bottom = re.match(r'([A-Z]+)(\d+)', bottom_right)
                    
                    if match_top and match_bottom:
                        top_col_letter = match_top.group(1)
                        top_row_num = int(match_top.group(2))
                        bottom_col_letter = match_bottom.group(1)
                        bottom_row_num = int(match_bottom.group(2))
                        
                        # Convertir letra de columna a número (A=1, B=2, etc.)
                        top_col_num = 0
                        for char in top_col_letter:
                            top_col_num = top_col_num * 26 + (ord(char) - ord('A') + 1)
                        
                        bottom_col_num = 0
                        for char in bottom_col_letter:
                            bottom_col_num = bottom_col_num * 26 + (ord(char) - ord('A') + 1)
                        
                        # Verificar si estamos dentro del rango fusionado
                        if (top_row_num <= row <= bottom_row_num and 
                            top_col_num <= col <= bottom_col_num):
                            # Solo escribir si es la celda principal (top-left)
                            if row == top_row_num and col == top_col_num:
                                cell.value = value
                            # Si no es la principal, no escribir (es parte del merge)
                            return
        
        # Si no está en un merge, escribir normalmente
        cell.value = value
    except Exception as e:
        logger.warning(f"Error al escribir en celda ({row}, {col}): {e}, intentando método directo")
        # Intentar escribir directamente en la celda si falla
        try:
            ws.cell(row=row, column=col).value = value
        except Exception as e2:
            # Si es un error de celda fusionada, ignorarlo silenciosamente
            if "read-only" not in str(e2).lower() and "merged" not in str(e2).lower():
                logger.error(f"Error crítico al escribir en celda ({row}, {col}): {e2}")




def _save_workbook_to_bytes(wb: Workbook) -> bytes:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _get_month_year() -> str:
    """Obtiene el mes y año en español"""
    now = datetime.now()
    months = [
        'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
        'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE'
    ]
    return f'{months[now.month - 1]} {now.year}'


def _get_jumper_category_color(tipo: str) -> Optional[str]:
    """Obtiene el color hexadecimal para una categoría de jumper según el tipo"""
    if not tipo:
        return None
    
    tipo_upper = tipo.upper().strip()
    
    # Mapeo de categorías a colores (mismos colores que en el frontend)
    color_map = {
        'FC-FC': 'FF2196F3',      # Colors.blue
        'FC-LC': 'FF3F51B5',      # Colors.indigo
        'FC-SC': 'FF673AB7',      # Colors.deepPurple
        'LC-FC': 'FF4CAF50',      # Colors.green
        'LC-LC': 'FFFF9800',      # Colors.orange
        'SC-FC': 'FF9C27B0',      # Colors.purple
        'SC-LC': 'FFF44336',      # Colors.red
        'SC-SC': 'FF009688',      # Colors.teal
    }
    
    # Buscar coincidencia exacta o parcial
    for category, color in color_map.items():
        if category in tipo_upper or tipo_upper in category:
            return color
    
    return None


def _apply_cell_style(cell, bold: bool = False, center: bool = True):
    """Aplica estilo a una celda"""
    if bold:
        cell.font = Font(bold=True)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Bordes delgados
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border


def _create_jumpers_excel(items: List[Dict[str, Any]]) -> Workbook:
    """Crea un archivo Excel para jumpers con el formato correcto"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Eliminar hoja por defecto si existe otra
    if "Sheet" in wb.sheetnames and ws.title != "Sheet":
        wb.remove(wb["Sheet"])
    
    # Título (fila 0, columna C - índice 2)
    title_cell = ws.cell(row=1, column=3, value=f'INVENTARIO JUMPERS {_get_month_year()}')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fila vacía (fila 1)
    ws.append([])
    
    # Encabezados (fila 2)
    headers = ['TIPO', 'TAMAÑO (metros)', 'CANTIDAD', 'RACK', 'CONTENEDOR']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        _apply_cell_style(cell, bold=True, center=True)
    
    # Datos (fila 3 en adelante)
    for item in items:
        row_data = [
            item.get("tipo", item.get("categoryName", "")),
            item.get("tamano", item.get("size", "")),
            item.get("cantidad", item.get("quantity", 0)),
            item.get("rack", ""),
            item.get("contenedor", item.get("container", ""))
        ]
        ws.append(row_data)
        
        # Aplicar estilo a datos
        row_num = ws.max_row
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col)
            _apply_cell_style(cell, bold=False, center=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25.0  # TIPO
    ws.column_dimensions['B'].width = 12.0  # TAMAÑO
    ws.column_dimensions['C'].width = 12.0  # CANTIDAD
    ws.column_dimensions['D'].width = 15.0  # RACK
    ws.column_dimensions['E'].width = 15.0  # CONTENEDOR
    
    return wb


def _create_computo_excel(items: List[Dict[str, Any]]) -> Workbook:
    """Crea un archivo Excel para inventarios de cómputo con todos los campos del esquema SQL"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario Cómputo"
    
    if "Sheet" in wb.sheetnames and ws.title != "Sheet":
        wb.remove(wb["Sheet"])
    
    # Título
    title_cell = ws.cell(row=1, column=1, value=f'INVENTARIO EQUIPO DE CÓMPUTO {_get_month_year()}')
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=21)
    
    ws.append([])
    
    # Encabezados completos basados en t_equipos_computo
    headers = [
        'INVENTARIO', 'EQUIPO PM', 'FECHA REGISTRO', 'TIPO EQUIPO', 'MARCA', 'MODELO', 
        'PROCESADOR', 'NÚMERO SERIE', 'DISCO DURO', 'MEMORIA', 
        'SISTEMA OPERATIVO', 'ETIQUETA SO', 'OFFICE INSTALADO', 'TIPO USO', 
        'NOMBRE EQUIPO DOMINIO', 'STATUS', 
        'UBICACIÓN FÍSICA', 'UBICACIÓN ADMINISTRATIVA', 
        'EMPLEADO ASIGNADO', 'EMPLEADO RESPONSABLE', 'OBSERVACIONES'
    ]
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        _apply_cell_style(cell, bold=True, center=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=10)
    
    # Datos
    for item in items:
        row_data = [
            item.get("inventario", item.get("inventario", "")),
            item.get("equipo_pm", item.get("equipo_pm", "")),
            item.get("fecha_registro", item.get("fecha_registro", "")),
            item.get("tipo_equipo", item.get("tipo_equipo", "")),
            item.get("marca", item.get("marca", "")),
            item.get("modelo", item.get("modelo", "")),
            item.get("procesador", item.get("procesador", "")),
            item.get("numero_serie", item.get("numero_serie", item.get("serie", ""))),
            item.get("disco_duro", item.get("disco_duro", "")),
            item.get("memoria", item.get("memoria", "")),
            item.get("sistema_operativo", item.get("sistema_operativo_instalado", "")),
            item.get("etiqueta_so", item.get("etiqueta_sistema_operativo", "")),
            item.get("office_instalado", item.get("office_instalado", "")),
            item.get("tipo_uso", item.get("tipo_uso", "")),
            item.get("nombre_dominio", item.get("nombre_equipo_dominio", "")),
            item.get("status", item.get("status", "ASIGNADO")),
            item.get("ubicacion_fisica", item.get("ubicacion_fisica", "")),
            item.get("ubicacion_admin", item.get("ubicacion_administrativa", "")),
            item.get("empleado_asignado", item.get("empleado_asignado", "")),
            item.get("empleado_responsable", item.get("empleado_responsable", "")),
            item.get("observaciones", item.get("observaciones", ""))
        ]
        ws.append(row_data)
        
        row_num = ws.max_row
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col)
            _apply_cell_style(cell, bold=False, center=False)
            # Alternar colores de fila para mejor legibilidad
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ajustar ancho de columnas
    column_widths = [15.0, 12.0, 12.0, 15.0, 12.0, 15.0, 15.0, 18.0, 12.0, 10.0, 
                     18.0, 12.0, 15.0, 12.0, 20.0, 12.0, 20.0, 20.0, 20.0, 20.0, 30.0]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Congelar paneles (fijar encabezados)
    ws.freeze_panes = 'A3'
    
    return wb


def _create_sdr_excel(items: List[Dict[str, Any]]) -> Workbook:
    """Crea un archivo Excel para formatos SDR"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    if "Sheet" in wb.sheetnames and ws.title != "Sheet":
        wb.remove(wb["Sheet"])
    
    # Título
    title_cell = ws.cell(row=1, column=2, value=f'INVENTARIO SDR {_get_month_year()}')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.append([])
    
    # Encabezados para SDR
    headers = ['CÓDIGO', 'DESCRIPCIÓN', 'CANTIDAD', 'UBICACIÓN', 'FECHA', 'OBSERVACIONES']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        _apply_cell_style(cell, bold=True, center=True)
    
    # Datos
    for item in items:
        row_data = [
            item.get("codigo", item.get("code", "")),
            item.get("descripcion", item.get("description", "")),
            item.get("cantidad", item.get("quantity", 0)),
            item.get("ubicacion", item.get("location", "")),
            item.get("fecha", item.get("date", "")),
            item.get("observaciones", item.get("notes", ""))
        ]
        ws.append(row_data)
        
        row_num = ws.max_row
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col)
            _apply_cell_style(cell, bold=False, center=True)
    
    # Ajustar ancho de columnas
    column_widths = [15.0, 30.0, 12.0, 15.0, 12.0, 25.0]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    return wb


@app.get("/", tags=["root"])
def root():
    return {
        "ok": True,
        "endpoints": [
            "/api/generate-jumpers-excel",
            "/api/generate-computo-excel",
            "/api/generate-sdr-excel",
            "/api/debug-last-file",
            "/health"
        ]
    }


@app.get("/health", tags=["health"])
def health():
    try:
        templates_status = {
            "jumpers": os.path.exists(TEMPLATE_PATH_JUMPERS),
            "computo": os.path.exists(TEMPLATE_PATH_COMPUTO),
            "sdr": os.path.exists(TEMPLATE_PATH_SDR)
        }
        return {"ok": True, "templates": templates_status}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.post("/api/generate-jumpers-excel")
async def generate_jumpers_excel(request: Request):
    payload = await request.json()
    items: List[Dict[str, Any]] = payload.get("items") or []
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items must be a non-empty list")

    try:
        # Intentar usar plantilla si existe
        if _ensure_template(TEMPLATE_PATH_JUMPERS):
            wb = openpyxl.load_workbook(TEMPLATE_PATH_JUMPERS)
            ws = wb.active
            
            # Los datos empiezan en la fila 5 según la plantilla
            # Columnas: B=TIPO, C=TAMAÑO, D=CANTIDAD, E=RACK, F=CONTENEDOR (o #)
            start_row = 5
            
            # Buscar la columna UBICACION en los encabezados PRIMERO
            # Buscar en varias filas por si cambia la estructura de la plantilla
            ubicacion_col = None
            for row_header in [4, 3, 2, 1]:  # Buscar en varias filas
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_header, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).upper().strip()
                        # Buscar variaciones: UBICACION, UBICACIÓN, UBIC, LOCATION
                        if "UBICACION" in cell_str or "UBICACIÓN" in cell_str or "UBIC" in cell_str:
                            ubicacion_col = col
                            logger.info(f"📍 Columna UBICACION encontrada en fila {row_header}, columna {col}")
                            break
                if ubicacion_col:
                    break
            
            # Si no se encuentra UBICACION, usar la columna E (5) como fallback
            # (normalmente es UBICACION después de TIPO, TAMAÑO, CANTIDAD)
            if ubicacion_col is None:
                ubicacion_col = 5
                logger.warning(f"⚠️ Columna UBICACION no encontrada, usando columna {ubicacion_col} como fallback")
            
            # Obtener formato de referencia de la fila 5
            # Incluir la columna UBICACION en el rango si está dentro de B-F, o extender el rango
            max_ref_col = max(7, ubicacion_col + 1)  # Asegurar que incluya UBICACION
            reference_cells = {}
            for col in range(2, max_ref_col):  # Columnas B hasta incluir UBICACION
                ref_cell = ws.cell(row=start_row, column=col)
                reference_cells[col] = {
                    'font': ref_cell.font.copy() if ref_cell.font else None,
                    'fill': ref_cell.fill.copy() if ref_cell.fill else None,
                    'border': ref_cell.border.copy() if ref_cell.border else None,
                    'alignment': ref_cell.alignment.copy() if ref_cell.alignment else None,
                    'number_format': ref_cell.number_format,
                }
            
            # Insertar datos empezando desde la fila 5
            for idx, item in enumerate(items, start=0):
                row = start_row + idx
                
                # Obtener el tipo para determinar el color
                tipo = item.get("tipo", item.get("categoryName", ""))
                tipo_color = _get_jumper_category_color(tipo)
                
                # Col B: TIPO
                _safe_set_cell_value(ws, row, 2, tipo)
                # Col C: TAMAÑO (metros)
                _safe_set_cell_value(ws, row, 3, item.get("tamano", item.get("size", "")))
                # Col D: CANTIDAD
                _safe_set_cell_value(ws, row, 4, item.get("cantidad", item.get("quantity", 0)))
                
                # Columna UBICACION: Formatear contenedores múltiples como R{rack}-{contenedor}
                # Solo se escribe en UBICACION, NO en columnas RACK/CONTENEDOR por separado
                contenedores = item.get("contenedores", [])
                ubicaciones = []
                
                if contenedores and len(contenedores) > 0:
                    # Formatear cada contenedor como R{rack}-{contenedor}
                    for cont in contenedores:
                        rack = str(cont.get("rack", "")).strip()
                        contenedor = str(cont.get("contenedor", "")).strip()
                        
                        if rack and contenedor:
                            # Extraer número del rack (ej: "1" de "Rack 1" o "1")
                            rack_num = rack
                            if "rack" in rack.lower():
                                # Si contiene "rack", extraer el número
                                match = re.search(r'\d+', rack)
                                if match:
                                    rack_num = match.group()
                            
                            ubicacion = f"R{rack_num}-{contenedor}"
                            ubicaciones.append(ubicacion)
                        elif contenedor:
                            # Si solo hay contenedor sin rack, solo mostrar el contenedor
                            ubicaciones.append(contenedor)
                
                # Si no hay contenedores múltiples, usar rack/contenedor antiguo como fallback
                if not ubicaciones:
                    rack = str(item.get("rack", "")).strip()
                    contenedor = str(item.get("contenedor", item.get("container", ""))).strip()
                    if rack and contenedor:
                        rack_num = rack
                        if "rack" in rack.lower():
                            match = re.search(r'\d+', rack)
                            if match:
                                rack_num = match.group()
                        ubicaciones.append(f"R{rack_num}-{contenedor}")
                    elif contenedor:
                        ubicaciones.append(contenedor)
                
                # Combinar todas las ubicaciones en una sola celda (separadas por comas)
                ubicacion_text = ", ".join(ubicaciones) if ubicaciones else ""
                _safe_set_cell_value(ws, row, ubicacion_col, ubicacion_text)
                
                # Aplicar formato de la fila 5 a cada celda
                for col in range(2, 7):
                    cell = ws.cell(row=row, column=col)
                    ref_format = reference_cells[col]
                    
                    if ref_format['font']:
                        cell.font = ref_format['font']
                    if ref_format['fill']:
                        # Para la columna TIPO (columna B, índice 2), aplicar color según categoría
                        if col == 2 and tipo_color:
                            cell.fill = PatternFill(start_color=tipo_color, end_color=tipo_color, fill_type="solid")
                        else:
                            cell.fill = ref_format['fill']
                    if ref_format['border']:
                        cell.border = ref_format['border']
                    if ref_format['alignment']:
                        cell.alignment = ref_format['alignment']
                    if ref_format['number_format']:
                        cell.number_format = ref_format['number_format']
                
                # Aplicar formato a la columna UBICACION también
                if ubicacion_col:
                    ubicacion_cell = ws.cell(row=row, column=ubicacion_col)
                    
                    # Si la columna UBICACION está en el rango de referencia (B-F), usar ese formato
                    if ubicacion_col in reference_cells:
                        ref_format = reference_cells[ubicacion_col]
                        if ref_format['font']:
                            ubicacion_cell.font = ref_format['font']
                        if ref_format['fill']:
                            ubicacion_cell.fill = ref_format['fill']
                        if ref_format['border']:
                            ubicacion_cell.border = ref_format['border']
                        if ref_format['alignment']:
                            ubicacion_cell.alignment = ref_format['alignment']
                        if ref_format['number_format']:
                            ubicacion_cell.number_format = ref_format['number_format']
                    else:
                        # Si está fuera del rango, obtener formato de la fila de referencia
                        if ubicacion_col <= ws.max_column:
                            ref_cell = ws.cell(row=start_row, column=ubicacion_col)
                            
                            if ref_cell.font:
                                ubicacion_cell.font = ref_cell.font.copy()
                            if ref_cell.fill:
                                ubicacion_cell.fill = ref_cell.fill.copy()
                            if ref_cell.border:
                                ubicacion_cell.border = ref_cell.border.copy()
                            if ref_cell.alignment:
                                ubicacion_cell.alignment = ref_cell.alignment.copy()
                            if ref_cell.number_format:
                                ubicacion_cell.number_format = ref_cell.number_format
        else:
            # Crear desde cero con formato correcto si no hay plantilla
            wb = _create_jumpers_excel(items)

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_jumpers_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except Exception as e:
        logger.exception("Error generating jumpers excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate-computo-excel")
async def generate_computo_excel(request: Request):
    payload = await request.json()
    items: List[Dict[str, Any]] = payload.get("items") or []
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items must be a non-empty list")

    try:
        # Usar plantilla si existe
        if _ensure_template(TEMPLATE_PATH_COMPUTO):
            logger.info(f"📄 Usando plantilla: {TEMPLATE_PATH_COMPUTO}")
            wb = openpyxl.load_workbook(TEMPLATE_PATH_COMPUTO)
            ws = wb.active
            
            # La inserción empieza en la fila 5 (celda A5)
            start_row = 5
            
            # Buscar la primera fila vacía desde la fila 5
            while ws.cell(row=start_row, column=1).value is not None:
                start_row += 1
            
            logger.info(f"📝 Escribiendo {len(items)} equipos desde la fila {start_row} (celda A{start_row})")
            
            # Obtener el formato de la fila 5 (fila de referencia)
            # La plantilla tiene 40 columnas según los encabezados
            reference_row = 5
            reference_cells = {}
            for col in range(1, 41):  # Columnas A-AN (40 columnas)
                ref_cell = ws.cell(row=reference_row, column=col)
                reference_cells[col] = {
                    'font': ref_cell.font.copy() if ref_cell.font else None,
                    'fill': ref_cell.fill.copy() if ref_cell.fill else None,
                    'border': ref_cell.border.copy() if ref_cell.border else None,
                    'alignment': ref_cell.alignment.copy() if ref_cell.alignment else None,
                    'number_format': ref_cell.number_format,
                }
            
            # Ordenar items por ID de menor a mayor
            def get_id_value(item):
                id_val = item.get("id")
                if id_val is None:
                    return 0
                try:
                    if isinstance(id_val, int):
                        return id_val
                    if isinstance(id_val, str):
                        return int(id_val) if id_val.isdigit() else 0
                    return int(id_val)
                except (ValueError, TypeError):
                    return 0
            
            sorted_items = sorted(items, key=get_id_value)
            
            # Escribir cada equipo/accesorio en una fila usando función segura y copiando formato
            for idx, item in enumerate(sorted_items, start=0):
                row = start_row + idx
                
                # Mapear campos según la plantilla (40 columnas)
                # Col A (1): ID
                _safe_set_cell_value(ws, row, 1, item.get("id", idx + 1))
                # Col B (2): INVENTARIO
                _safe_set_cell_value(ws, row, 2, item.get("inventario", ""))
                # Col C (3): EQUIPO PM
                _safe_set_cell_value(ws, row, 3, item.get("equipo_pm", ""))
                # Col D (4): FECHA REGISTRO
                _safe_set_cell_value(ws, row, 4, item.get("fecha_registro", ""))
                # Col E (5): TIPO DE EQUIPO
                _safe_set_cell_value(ws, row, 5, item.get("tipo_equipo", ""))
                # Col F (6): MARCA
                _safe_set_cell_value(ws, row, 6, item.get("marca", ""))
                # Col G (7): MODELO
                _safe_set_cell_value(ws, row, 7, item.get("modelo", ""))
                # Col H (8): PROCESADOR
                _safe_set_cell_value(ws, row, 8, item.get("procesador", ""))
                # Col I (9): NUMERO DE SERIE
                _safe_set_cell_value(ws, row, 9, item.get("numero_serie", ""))
                # Col J (10): DISCO DURO
                _safe_set_cell_value(ws, row, 10, item.get("disco_duro", ""))
                # Col K (11): MEMORIA
                _safe_set_cell_value(ws, row, 11, item.get("memoria", ""))
                # Col L (12): SISTEMA OPERATIVO INSTALADO
                _safe_set_cell_value(ws, row, 12, item.get("sistema_operativo_instalado", item.get("sistema_operativo", "")))
                # Col M (13): ETIQUETA DE SISTEMA OPERATIVO
                _safe_set_cell_value(ws, row, 13, item.get("etiqueta_sistema_operativo", ""))
                # Col N (14): OFFICE INSTALADO
                _safe_set_cell_value(ws, row, 14, item.get("office_instalado", ""))
                # Col O (15): DIRECCIÓN FISICA DEL EQUIPO
                _safe_set_cell_value(ws, row, 15, item.get("direccion_fisica", item.get("ubicacion_fisica", "")))
                # Col P (16): ESTADO
                _safe_set_cell_value(ws, row, 16, item.get("estado", ""))
                # Col Q (17): CIUDAD
                _safe_set_cell_value(ws, row, 17, item.get("ciudad", ""))
                # Col R (18): TIPO DE EDIFICIO
                _safe_set_cell_value(ws, row, 18, item.get("tipo_edificio", ""))
                # Col S (19): NOMBRE DEL EDIFICIO
                _safe_set_cell_value(ws, row, 19, item.get("nombre_edificio", ""))
                # Col T (20): TIPO DE USO
                _safe_set_cell_value(ws, row, 20, item.get("tipo_uso", ""))
                # Col U (21): NOMBRE DEL EQUIPO EN DOMINIO
                _safe_set_cell_value(ws, row, 21, item.get("nombre_equipo_dominio", ""))
                # Col V (22): STATUS
                _safe_set_cell_value(ws, row, 22, item.get("status", ""))
                # Col W (23): DIRECCIÓN ADMINISTRATIVA
                _safe_set_cell_value(ws, row, 23, item.get("direccion_administrativa", ""))
                # Col X (24): SUBDIRECCIÓN
                _safe_set_cell_value(ws, row, 24, item.get("subdireccion", ""))
                # Col Y (25): GERENCIA
                _safe_set_cell_value(ws, row, 25, item.get("gerencia", ""))
                # Col Z (26): EXPEDIENTE (Usuario Responsable) - INTERCAMBIO: La plantilla tiene Responsable primero
                _safe_set_cell_value(ws, row, 26, item.get("expediente_responsable", ""))
                # Col AA (27): NOMBRE COMPLETO (Usuario Responsable)
                _safe_set_cell_value(ws, row, 27, item.get("nombre_completo_responsable", ""))
                # Col AB (28): APELLIDO PATERNO (Usuario Responsable)
                _safe_set_cell_value(ws, row, 28, item.get("apellido_paterno_responsable", ""))
                # Col AC (29): APELLIDO MATERNO (Usuario Responsable)
                _safe_set_cell_value(ws, row, 29, item.get("apellido_materno_responsable", ""))
                # Col AD (30): NOMBRE (Usuario Responsable)
                _safe_set_cell_value(ws, row, 30, item.get("nombre_responsable", ""))
                # Col AE (31): EMPRESA (Usuario Responsable)
                _safe_set_cell_value(ws, row, 31, item.get("empresa_responsable", ""))
                # Col AF (32): PUESTO (Usuario Responsable)
                _safe_set_cell_value(ws, row, 32, item.get("puesto_responsable", ""))
                # Col AG (33): EXPEDIENTE (Usuario Final) - INTERCAMBIO: La plantilla tiene Final después
                _safe_set_cell_value(ws, row, 33, item.get("expediente_final", ""))
                # Col AH (34): NOMBRE COMPLETO (Usuario Final)
                _safe_set_cell_value(ws, row, 34, item.get("nombre_completo_final", ""))
                # Col AI (35): APELLIDO PATERNO (Usuario Final)
                _safe_set_cell_value(ws, row, 35, item.get("apellido_paterno_final", ""))
                # Col AJ (36): APELLIDO MATERNO (Usuario Final)
                _safe_set_cell_value(ws, row, 36, item.get("apellido_materno_final", ""))
                # Col AK (37): NOMBRE (Usuario Final)
                _safe_set_cell_value(ws, row, 37, item.get("nombre_final", ""))
                # Col AL (38): EMPRESA (Usuario Final)
                _safe_set_cell_value(ws, row, 38, item.get("empresa_final", ""))
                # Col AM (39): PUESTO (Usuario Final)
                _safe_set_cell_value(ws, row, 39, item.get("puesto_final", ""))
                # Col AN (40): OBSERVACIONES
                _safe_set_cell_value(ws, row, 40, item.get("observaciones", ""))
                
                # Aplicar formato de la fila 5 a cada celda
                for col in range(1, 41):
                    cell = ws.cell(row=row, column=col)
                    ref_format = reference_cells[col]
                    
                    if ref_format['font']:
                        cell.font = ref_format['font']
                    if ref_format['fill']:
                        cell.fill = ref_format['fill']
                    if ref_format['border']:
                        cell.border = ref_format['border']
                    if ref_format['alignment']:
                        cell.alignment = ref_format['alignment']
                    if ref_format['number_format']:
                        cell.number_format = ref_format['number_format']
            
            # Detectar grupos de filas con el mismo ID y EQUIPO PM para combinar celdas
            # Columna A (ID) y Columna C (EQUIPO PM)
            # Agrupar por (ID, EQUIPO_PM) como tupla
            groups = {}
            for idx, item in enumerate(sorted_items, start=0):
                row = start_row + idx
                item_id = item.get("id")
                item_equipo_pm = item.get("equipo_pm", "")
                group_key = (item_id, item_equipo_pm)
                
                if group_key not in groups:
                    groups[group_key] = {'start_row': row, 'end_row': row}
                else:
                    groups[group_key]['end_row'] = row
            
            # Combinar celdas para cada grupo
            for group_key, group_info in groups.items():
                start_row_group = group_info['start_row']
                end_row_group = group_info['end_row']
                
                # Solo combinar si hay más de una fila en el grupo
                if end_row_group > start_row_group:
                    # Combinar celdas de ID (columna A) para este grupo
                    try:
                        ws.merge_cells(start_row=start_row_group, start_column=1, end_row=end_row_group, end_column=1)
                        # Centrar el texto en la celda combinada
                        merged_cell = ws.cell(row=start_row_group, column=1)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        logger.info(f"✅ Celdas de ID combinadas: filas {start_row_group}-{end_row_group} (ID: {group_key[0]})")
                    except Exception as e:
                        logger.warning(f"⚠️ Error al combinar celdas de ID (filas {start_row_group}-{end_row_group}): {e}")
                    
                    # Combinar celdas de EQUIPO PM (columna C) para este grupo
                    try:
                        ws.merge_cells(start_row=start_row_group, start_column=3, end_row=end_row_group, end_column=3)
                        # Centrar el texto en la celda combinada
                        merged_cell = ws.cell(row=start_row_group, column=3)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                        logger.info(f"✅ Celdas de EQUIPO PM combinadas: filas {start_row_group}-{end_row_group} (EQUIPO PM: {group_key[1]})")
                    except Exception as e:
                        logger.warning(f"⚠️ Error al combinar celdas de EQUIPO PM (filas {start_row_group}-{end_row_group}): {e}")
        else:
            # Crear desde cero con formato correcto
            wb = _create_computo_excel(items)

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_computo_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except Exception as e:
        logger.exception("Error generating computo excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate-sdr-excel")
async def generate_sdr_excel(request: Request):
    payload = await request.json()
    items: List[Dict[str, Any]] = payload.get("items") or []
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items must be a non-empty list")

    try:
        # Intentar usar plantilla si existe, sino crear desde cero
        if _ensure_template(TEMPLATE_PATH_SDR):
            wb = openpyxl.load_workbook(TEMPLATE_PATH_SDR)
            ws = wb.active
            
            # Tomar el primer item (ya que es un formulario único, no una lista de items)
            item = items[0] if items else {}
            
            # Mapear campos a las filas correspondientes de la plantilla
            # Las columnas B y C están combinadas, así que escribimos en B
            # Datos de Falla de aviso
            ws.cell(row=9, column=2, value=item.get("fecha", item.get("date", "")))  # Fecha
            ws.cell(row=10, column=2, value=item.get("descripcion_aviso", item.get("descripcion_del_aviso", "")))  # Descripción del Aviso
            ws.cell(row=11, column=2, value=item.get("grupo_planificador", ""))  # Grupo planificador
            ws.cell(row=12, column=2, value=item.get("puesto_trabajo_responsable", ""))  # Puesto de trabajo responsable
            ws.cell(row=13, column=2, value=item.get("autor_aviso", ""))  # Autor de aviso
            ws.cell(row=14, column=2, value=item.get("motivo_intervencion", ""))  # Motivo de intervención
            ws.cell(row=15, column=2, value=item.get("modelo_dano", item.get("modelo_del_dano", "")))  # Modelo del Daño
            ws.cell(row=16, column=2, value=item.get("causa_averia", ""))  # Causa de la avería
            ws.cell(row=17, column=2, value=item.get("repercusion_funcionamiento", ""))  # Repercusión en el funcionamiento
            ws.cell(row=18, column=2, value=item.get("estado_instalacion", ""))  # Estado de la Instalación
            ws.cell(row=19, column=2, value=item.get("motivo_intervencion_afectacion", ""))  # Motivo de Intervención (AFECTACION)
            ws.cell(row=21, column=2, value=item.get("atencion_dano", ""))  # Atención del Daño
            ws.cell(row=22, column=2, value=item.get("prioridad", ""))  # Prioridad
            
            # Lugar del Daño
            ws.cell(row=25, column=2, value=item.get("centro_emplazamiento", ""))  # Centro Emplazamiento
            ws.cell(row=26, column=2, value=item.get("area_empresa", ""))  # Área de empresa
            ws.cell(row=27, column=2, value=item.get("puesto_trabajo_emplazamiento", ""))  # Puesto trabajo de emplazamiento
            ws.cell(row=28, column=2, value=item.get("division", ""))  # División
            ws.cell(row=29, column=2, value=item.get("estado_instalacion_lugar", ""))  # Estado de Instalación
            ws.cell(row=30, column=2, value=item.get("datos_disponibles", ""))  # Datos disponibles
            ws.cell(row=32, column=2, value=item.get("emplazamiento_1", item.get("emplazamiento", "")))  # Emplazamiento (primera ocurrencia)
            ws.cell(row=33, column=2, value=item.get("emplazamiento_2", item.get("emplazamiento", "")))  # Emplazamiento (segunda ocurrencia)
            ws.cell(row=34, column=2, value=item.get("local", ""))  # Local
            ws.cell(row=35, column=2, value=item.get("campo_clasificacion", ""))  # Campo de clasificación
            
            # Datos de la unidad Dañada
            ws.cell(row=38, column=2, value=item.get("tipo_unidad_danada", ""))  # Tipo
            ws.cell(row=39, column=2, value=item.get("no_serie_unidad_danada", ""))  # No de serie
            
            # Datos de la unidad que se montó
            ws.cell(row=42, column=2, value=item.get("tipo_unidad_montada", ""))  # Tipo
            ws.cell(row=43, column=2, value=item.get("no_serie_unidad_montada", ""))  # No de serie
            
        else:
            # Crear desde cero con formato correcto
            wb = _create_sdr_excel(items)

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"solicitud_sdr_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except Exception as e:
        logger.exception("Error generating SDR excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate-sicor-excel")
async def generate_sicor_excel(request: Request):
    payload = await request.json()
    items: List[Dict[str, Any]] = payload.get("items") or []
    if not isinstance(items, list) or len(items) == 0:
        raise HTTPException(status_code=400, detail="items must be a non-empty list")

    try:
        # Usar plantilla si existe
        if _ensure_template(TEMPLATE_PATH_SICOR):
            logger.info(f"📄 Usando plantilla: {TEMPLATE_PATH_SICOR}")
            wb = openpyxl.load_workbook(TEMPLATE_PATH_SICOR)
            ws = wb.active
            
            # Actualizar la fecha en el encabezado (fila 2, celda C2 que está en merged cell C2:H2)
            try:
                header_cell = ws.cell(row=2, column=3)  # Columna C, fila 2
                header_text = str(header_cell.value) if header_cell.value else ""
                
                # Obtener fecha actual en formato DD/MM/YYYY
                now = datetime.now()
                fecha_actual = now.strftime("%d/%m/%Y")
                
                # Reemplazar la fecha al final del texto (formato DD/MM/YYYY)
                # Buscar patrón de fecha al final: " - DD/MM/YYYY" o " - DD/MM/YYYY" al final
                # Mantener todo el texto antes de la fecha
                pattern = r'\s*-\s*\d{2}/\d{2}/\d{4}\s*$'
                if re.search(pattern, header_text):
                    # Reemplazar la fecha al final
                    nuevo_texto = re.sub(pattern, f' - {fecha_actual}', header_text)
                else:
                    # Si no hay fecha al final, agregarla
                    nuevo_texto = f"{header_text.rstrip()} - {fecha_actual}"
                
                # Actualizar la celda con el nuevo texto
                header_cell.value = nuevo_texto
                logger.info(f"📅 Fecha actualizada en encabezado: {fecha_actual}")
            except Exception as e:
                logger.warning(f"⚠️ No se pudo actualizar la fecha en el encabezado: {e}")
            
            # Los datos empiezan en la fila 5, columna B (columna 2)
            start_row = 5
            start_col = 2  # Columna B
            
            # Buscar la primera fila vacía desde la fila 5
            while ws.cell(row=start_row, column=start_col).value is not None:
                start_row += 1
            
            logger.info(f"📝 Escribiendo {len(items)} tarjetas desde la fila {start_row}, columna {start_col}")
            
            # Obtener el formato de la fila 5 (fila de referencia) si existe
            reference_row = 5
            reference_cells = {}
            for col in range(start_col, start_col + 7):  # 7 columnas: B-H
                ref_cell = ws.cell(row=reference_row, column=col)
                reference_cells[col] = {
                    'font': ref_cell.font.copy() if ref_cell.font else None,
                    'fill': ref_cell.fill.copy() if ref_cell.fill else None,
                    'border': ref_cell.border.copy() if ref_cell.border else None,
                    'alignment': ref_cell.alignment.copy() if ref_cell.alignment else None,
                    'number_format': ref_cell.number_format,
                }
            
            # Escribir cada tarjeta en una fila usando función segura y copiando formato
            for idx, item in enumerate(items, start=0):
                row = start_row + idx
                en_stock = str(item.get("en_stock", "SI")).upper().strip()
                is_no_stock = en_stock == "NO"
                
                # Mapear campos según la plantilla (empezando en columna B)
                # Col B (2): EN STOCK -> en_stock
                _safe_set_cell_value(ws, row, 2, en_stock)
                # Col C (3): No. -> numero
                _safe_set_cell_value(ws, row, 3, item.get("numero", ""))
                # Col D (4): CODIGO -> codigo
                _safe_set_cell_value(ws, row, 4, item.get("codigo", ""))
                # Col E (5): SERIE -> serie
                _safe_set_cell_value(ws, row, 5, item.get("serie", ""))
                # Col F (6): MARCA -> marca
                _safe_set_cell_value(ws, row, 6, item.get("marca", ""))
                # Col G (7): POSICION -> posicion
                _safe_set_cell_value(ws, row, 7, item.get("posicion", ""))
                # Col H (8): COMENTARIOS -> comentarios
                _safe_set_cell_value(ws, row, 8, item.get("comentarios", ""))
                
                # Aplicar formato de la fila 5 a cada celda
                for col in range(start_col, start_col + 7):
                    cell = ws.cell(row=row, column=col)
                    ref_format = reference_cells.get(col, {})
                    
                    # Si NO está en stock, aplicar fondo rojo a toda la fila
                    if is_no_stock:
                        # Fondo rojo para toda la fila
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        
                        # Texto blanco para columnas B (en_stock) y C (numero)
                        if col == 2 or col == 3:  # Columnas B y C
                            ref_font = ref_format.get('font')
                            cell.font = Font(
                                color="FFFFFF",  # Blanco
                                bold=ref_font.bold if ref_font else False,
                                size=ref_font.size if ref_font else 11
                            )
                        else:
                            # Para otras columnas, mantener el formato original pero con fondo rojo
                            if ref_format.get('font'):
                                ref_font = ref_format['font']
                                font_copy = Font(
                                    color=ref_font.color if ref_font.color else "000000",
                                    bold=ref_font.bold if ref_font.bold is not None else False,
                                    size=ref_font.size if ref_font.size else 11
                                )
                                cell.font = font_copy
                    else:
                        # Si está en stock, aplicar formato normal
                        if ref_format.get('font'):
                            cell.font = ref_format['font']
                        if ref_format.get('fill'):
                            cell.fill = ref_format['fill']
                        
                        # Columna D (CODIGO) con fondo azul #558ED5 cuando está en stock
                        if col == 4:  # Columna D (CODIGO)
                            cell.fill = PatternFill(start_color="558ED5", end_color="558ED5", fill_type="solid")
                    
                    # Aplicar otros formatos (border, alignment, number_format) siempre
                    if ref_format.get('border'):
                        cell.border = ref_format['border']
                    if ref_format.get('alignment'):
                        cell.alignment = ref_format['alignment']
                    if ref_format.get('number_format'):
                        cell.number_format = ref_format['number_format']
        else:
            # Si no hay plantilla, crear estructura básica
            logger.warning("No se encontró plantilla SICOR, creando estructura básica")
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario SICOR"
            
            # Título
            title_cell = ws.cell(row=1, column=2, value=f'INVENTARIO SICOR {_get_month_year()}')
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Encabezados (fila 4)
            headers = ['EN STOCK', 'No.', 'CODIGO', 'SERIE', 'MARCA', 'POSICION', 'COMENTARIOS']
            for idx, header in enumerate(headers, start=2):
                cell = ws.cell(row=4, column=idx, value=header)
                _apply_cell_style(cell, bold=True, center=True)
            
            # Datos (fila 5 en adelante)
            for idx, item in enumerate(items, start=0):
                row = 5 + idx
                en_stock = str(item.get("en_stock", "SI")).upper().strip()
                is_no_stock = en_stock == "NO"
                
                ws.cell(row=row, column=2, value=en_stock)
                ws.cell(row=row, column=3, value=item.get("numero", ""))
                ws.cell(row=row, column=4, value=item.get("codigo", ""))
                ws.cell(row=row, column=5, value=item.get("serie", ""))
                ws.cell(row=row, column=6, value=item.get("marca", ""))
                ws.cell(row=row, column=7, value=item.get("posicion", ""))
                ws.cell(row=row, column=8, value=item.get("comentarios", ""))
                
                # Aplicar estilo
                for col in range(2, 9):
                    cell = ws.cell(row=row, column=col)
                    
                    # Si NO está en stock, aplicar fondo rojo a toda la fila
                    if is_no_stock:
                        # Fondo rojo para toda la fila
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        
                        # Texto blanco para columnas B (en_stock) y C (numero)
                        if col == 2 or col == 3:  # Columnas B y C
                            cell.font = Font(color="FFFFFF", bold=False, size=11)
                        else:
                            # Para otras columnas, mantener texto normal pero con fondo rojo
                            cell.font = Font(color="000000", bold=False, size=11)
                    else:
                        # Si está en stock, aplicar estilo normal
                        _apply_cell_style(cell, bold=False, center=True)
                        
                        # Columna D (CODIGO) con fondo azul #558ED5 cuando está en stock
                        if col == 4:  # Columna D (CODIGO)
                            cell.fill = PatternFill(start_color="558ED5", end_color="558ED5", fill_type="solid")

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_sicor_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except Exception as e:
        logger.exception("Error generating SICOR excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/generate-bitacora-excel")
async def generate_bitacora_excel(request: Request):
    payload = await request.json()
    
    # Nuevo formato: years_data es una lista de objetos con 'year' e 'items'
    years_data: List[Dict[str, Any]] = payload.get("years_data") or []
    
    # Compatibilidad con formato antiguo (un solo año)
    if not years_data:
        items: List[Dict[str, Any]] = payload.get("items") or []
        year = payload.get("year", datetime.now().year)
        if items:
            years_data = [{"year": year, "items": items}]
    
    if not isinstance(years_data, list) or len(years_data) == 0:
        raise HTTPException(status_code=400, detail="years_data must be a non-empty list")
    
    # Ordenar años de forma ascendente
    years_data.sort(key=lambda x: x.get("year", 0))

    try:
        # Crear un nuevo workbook
        wb = Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # Usar plantilla si existe - CARGAR SOLO UNA VEZ para optimizar
        template_exists = _ensure_template(TEMPLATE_PATH_BITACORA)
        template_wb = None
        template_ws = None
        template_merged_ranges = None
        template_column_widths = None
        template_row_heights = None
        
        if template_exists:
            logger.info(f"📄 Cargando plantilla una vez: {TEMPLATE_PATH_BITACORA}")
            template_wb = openpyxl.load_workbook(TEMPLATE_PATH_BITACORA)
            template_ws = template_wb.active
            template_merged_ranges = list(template_ws.merged_cells.ranges)
            template_column_widths = {col: template_ws.column_dimensions[col].width for col in template_ws.column_dimensions}
            template_row_heights = {row: template_ws.row_dimensions[row].height for row in template_ws.row_dimensions}
        
        # Procesar cada año
        total_years = len(years_data)
        for idx, year_data in enumerate(years_data, start=1):
            year = year_data.get("year")
            items: List[Dict[str, Any]] = year_data.get("items") or []
            
            if not items:
                logger.warning(f"⚠️ No hay items para el año {year}, saltando...")
                continue
            
            logger.info(f"📝 [{idx}/{total_years}] Procesando año {year} con {len(items)} registros")
            
            # Crear o copiar hoja para este año
            if template_exists and template_ws:
                # Crear nueva hoja con el nombre del año
                ws = wb.create_sheet(title=str(year))
                
                # Copiar todas las celdas de la plantilla (reutilizando la plantilla cargada)
                for row in template_ws.iter_rows():
                    for cell in row:
                        new_cell = ws.cell(row=cell.row, column=cell.column)
                        new_cell.value = cell.value
                        if cell.has_style:
                            new_cell.font = cell.font.copy() if cell.font else None
                            new_cell.fill = cell.fill.copy() if cell.fill else None
                            new_cell.border = cell.border.copy() if cell.border else None
                            new_cell.alignment = cell.alignment.copy() if cell.alignment else None
                            new_cell.number_format = cell.number_format
                
                # Copiar merged cells (reutilizando los rangos guardados)
                for merged_range in template_merged_ranges:
                    ws.merge_cells(str(merged_range))
                
                # Copiar anchos de columna (reutilizando los anchos guardados)
                for col, width in template_column_widths.items():
                    ws.column_dimensions[col].width = width
                
                # Copiar altos de fila (reutilizando los altos guardados)
                for row, height in template_row_heights.items():
                    ws.row_dimensions[row].height = height
            else:
                # Crear hoja nueva sin plantilla
                ws = wb.create_sheet(title=str(year))
            
            # Actualizar la fecha en el encabezado si existe (similar a SICOR)
            try:
                # Buscar celda con fecha en las primeras filas
                for row in range(1, 5):
                    for col in range(1, 10):
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            cell_text = str(cell.value)
                            # Si contiene "fecha" o un patrón de fecha, actualizar
                            if "fecha" in cell_text.lower() or re.search(r'\d{2}/\d{2}/\d{4}', cell_text):
                                now = datetime.now()
                                fecha_actual = now.strftime("%d/%m/%Y")
                                pattern = r'\s*-\s*\d{2}/\d{2}/\d{4}\s*$'
                                if re.search(pattern, cell_text):
                                    nuevo_texto = re.sub(pattern, f' - {fecha_actual}', cell_text)
                                else:
                                    nuevo_texto = f"{cell_text.rstrip()} - {fecha_actual}"
                                cell.value = nuevo_texto
                                logger.info(f"📅 Fecha actualizada en encabezado (año {year}): {fecha_actual}")
                                break
            except Exception as e:
                logger.warning(f"⚠️ No se pudo actualizar la fecha en el encabezado (año {year}): {e}")
            
            # Los datos empiezan en B4 (fila 4, columna B = columna 2)
            start_row = 4
            start_col = 2  # Columna B
            
            # Buscar la primera fila vacía desde la fila 4 (B4)
            # Si B4 ya tiene datos, buscar la siguiente fila vacía
            while ws.cell(row=start_row, column=start_col).value is not None:
                start_row += 1
            
            logger.info(f"📝 Escribiendo {len(items)} registros de bitácora (año {year}) desde la fila {start_row}, columna B")
            
            # Obtener el formato de la fila 4 (B4) como referencia si existe
            reference_row = 4
            reference_cells = {}
            # 13 columnas empezando desde B: Consecutivo, Fecha, Técnico, Tarjeta, Código, Serie, Folio, Envía, Recibe, Guía, Anexos, COBO (INCIDENTE), Observaciones
            for col in range(start_col, start_col + 13):
                ref_cell = ws.cell(row=reference_row, column=col)
                reference_cells[col] = {
                    'font': ref_cell.font.copy() if ref_cell.font else None,
                    'fill': ref_cell.fill.copy() if ref_cell.fill else None,
                    'border': ref_cell.border.copy() if ref_cell.border else None,
                    'alignment': ref_cell.alignment.copy() if ref_cell.alignment else None,
                    'number_format': ref_cell.number_format,
                }
            
            # Escribir cada registro de bitácora en una fila empezando desde B4
            for idx, item in enumerate(items, start=0):
                row = start_row + idx
                
                # Mapear campos según la plantilla empezando desde columna B (2)
                # Columna B (2): Consecutivo
                _safe_set_cell_value(ws, row, 2, item.get("consecutivo", ""))
                # Columna C (3): Fecha
                fecha_str = item.get("fecha", "")
                if fecha_str:
                    try:
                        # Convertir de YYYY-MM-DD a DD/MM/YYYY si es necesario
                        fecha_date = datetime.strptime(fecha_str, "%Y-%m-%d")
                        fecha_formateada = fecha_date.strftime("%d/%m/%Y")
                        _safe_set_cell_value(ws, row, 3, fecha_formateada)
                    except:
                        _safe_set_cell_value(ws, row, 3, fecha_str)
                else:
                    _safe_set_cell_value(ws, row, 3, "")
                # Columna D (4): Técnico
                _safe_set_cell_value(ws, row, 4, item.get("tecnico", ""))
                # Columna E (5): Tarjeta
                _safe_set_cell_value(ws, row, 5, item.get("tarjeta", ""))
                # Columna F (6): Código
                _safe_set_cell_value(ws, row, 6, item.get("codigo", ""))
                # Columna G (7): Serie
                _safe_set_cell_value(ws, row, 7, item.get("serie", ""))
                # Columna H (8): Folio
                _safe_set_cell_value(ws, row, 8, item.get("folio", ""))
                # Columna I (9): Envía
                _safe_set_cell_value(ws, row, 9, item.get("envia", ""))
                # Columna J (10): Recibe
                _safe_set_cell_value(ws, row, 10, item.get("recibe", ""))
                # Columna K (11): Guía
                _safe_set_cell_value(ws, row, 11, item.get("guia", ""))
                # Columna L (12): Anexos
                _safe_set_cell_value(ws, row, 12, item.get("anexos", ""))
                # Columna M (13): COBO (en la plantilla se llama "INCIDENTE")
                _safe_set_cell_value(ws, row, 13, item.get("cobo", ""))
                # Columna N (14): Observaciones (última columna)
                _safe_set_cell_value(ws, row, 14, item.get("observaciones", ""))
                
                # Aplicar formato de la fila de referencia (B4) a cada celda
                for col in range(start_col, start_col + 13):
                    cell = ws.cell(row=row, column=col)
                    ref_format = reference_cells.get(col, {})
                    
                    if ref_format.get('font'):
                        cell.font = ref_format['font']
                    if ref_format.get('fill'):
                        cell.fill = ref_format['fill']
                    if ref_format.get('border'):
                        cell.border = ref_format['border']
                    if ref_format.get('alignment'):
                        cell.alignment = ref_format['alignment']
                    if ref_format.get('number_format'):
                        cell.number_format = ref_format['number_format']
            
            # Si no hay plantilla, crear estructura básica para esta hoja (solo encabezados)
            if not template_exists:
                # Título
                title_cell = ws.cell(row=1, column=1, value=f'BITÁCORA DE ENVÍOS - AÑO {year}')
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Encabezados (fila 3, empezando desde columna B)
                headers = ['Consecutivo', 'Fecha', 'Técnico', 'Tarjeta', 'Código', 'Serie', 'Folio', 
                          'Envía', 'Recibe', 'Guía', 'Anexos', 'INCIDENTE', 'Observaciones']
                for col, header in enumerate(headers, start=2):  # Empezar desde columna B (2)
                    cell = ws.cell(row=3, column=col, value=header)
                    _apply_cell_style(cell, bold=True, center=True)
        
        # Las hojas ya están ordenadas porque years_data está ordenado
        # Pero por si acaso, reordenarlas explícitamente
        # Crear un diccionario con las hojas
        sheets_dict = {ws.title: ws for ws in wb.worksheets}
        sorted_sheet_names = sorted(sheets_dict.keys(), key=lambda x: int(x) if x.isdigit() else 9999)
        
        # Reordenar las hojas moviendo cada una a su posición correcta
        for i, sheet_name in enumerate(sorted_sheet_names):
            if i == 0:
                continue  # La primera hoja ya está en su lugar
            sheet = sheets_dict[sheet_name]
            current_index = wb.index(sheet)
            target_index = i
            if current_index != target_index:
                # Mover la hoja a la posición correcta
                wb.move_sheet(sheet, offset=target_index - current_index)

        file_bytes = _save_workbook_to_bytes(wb)
        if not file_bytes:
            raise RuntimeError("Generated file is empty")

        global LAST_GENERATED_FILE_CONTENT, LAST_GENERATED_FILENAME
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        # Generar nombre de archivo con los años exportados
        years_list = sorted([str(yd.get("year", "")) for yd in years_data])
        years_str = "_".join(years_list)
        filename = f"bitacora_envio_{years_str}_{timestamp}.xlsx"
        LAST_GENERATED_FILE_CONTENT = file_bytes
        LAST_GENERATED_FILENAME = filename
        
        logger.info(f"📊 Archivo generado con {len(wb.worksheets)} hoja(s): {[ws.title for ws in wb.worksheets]}")

        logger.info(f"📦 Tamaño del archivo generado: {len(file_bytes)} bytes")

        return Response(content=file_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})

    except Exception as e:
        logger.exception("Error generating bitacora excel")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/debug-last-file")
def debug_last_file():
    if not LAST_GENERATED_FILE_CONTENT or not LAST_GENERATED_FILENAME:
        raise HTTPException(status_code=404, detail="No generated file in memory")

    tmp_path = os.path.join("/tmp", LAST_GENERATED_FILENAME)
    try:
        with open(tmp_path, "wb") as f:
            f.write(LAST_GENERATED_FILE_CONTENT)
        size = os.path.getsize(tmp_path)
        return {"ok": True, "path": tmp_path, "size": size}
    except Exception as e:
        logger.exception("Failed to write debug file")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("excel_generator_service.main:app", host="0.0.0.0", port=8001, reload=True)
